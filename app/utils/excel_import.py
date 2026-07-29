# -*- coding: utf-8 -*-
"""
Excel'dan savollarni o'qib, Question qatorlariga aylantiradi. Bu -- asosiy
kiritish yo'li admin panel (rasm/formula/jadval uchun), lekin oddiy
matnli savollarni tezkor qo'shish uchun bu ham ochiq qoladi.
"""
import openpyxl

REQUIRED_COLUMNS = ["T/r", "Fan", "Savol", "A", "B", "C", "D", "To'g'ri javob", "Ball"]
LETTERS = ("A", "B", "C", "D")


class ExcelImportError(Exception):
    pass


def parse_excel(file_obj, sheet_name="Savollar"):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    header = [c.value for c in ws[1]]
    col_idx = {}
    for name in REQUIRED_COLUMNS:
        if name not in header:
            raise ExcelImportError(f"Ustun topilmadi: '{name}'")
        col_idx[name] = header.index(name)

    rows = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        def cell(name):
            return row[col_idx[name]]

        savol, fan, togri, ball = cell("Savol"), cell("Fan"), cell("To'g'ri javob"), cell("Ball")
        variantlar = [cell("A"), cell("B"), cell("C"), cell("D")]

        if not savol or not fan:
            raise ExcelImportError(f"{row_num}-qator: 'Fan' yoki 'Savol' bo'sh")
        if not togri or str(togri).strip().upper() not in LETTERS:
            raise ExcelImportError(f"{row_num}-qator: 'To'g'ri javob' faqat A/B/C/D bo'lishi kerak")
        if any(v is None or str(v).strip() == "" for v in variantlar):
            raise ExcelImportError(f"{row_num}-qator: A/B/C/D variantlaridan biri bo'sh")
        try:
            ball_val = float(ball)
        except (TypeError, ValueError):
            raise ExcelImportError(f"{row_num}-qator: 'Ball' raqam bo'lishi kerak")

        rows.append({
            "fan": str(fan).strip(),
            "ball": ball_val,
            "savol_html": str(savol).strip(),
            "variant_a_html": str(variantlar[0]).strip(),
            "variant_b_html": str(variantlar[1]).strip(),
            "variant_c_html": str(variantlar[2]).strip(),
            "variant_d_html": str(variantlar[3]).strip(),
            "togri_javob": str(togri).strip().upper(),
        })

    if not rows:
        raise ExcelImportError("Faylda birorta ham to'liq savol topilmadi")
    return rows
