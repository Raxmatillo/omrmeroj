# -*- coding: utf-8 -*-
"""
excel_loader.py (booklet generator, savol_id-based) va excel_import.py
(admin panel Question CRUD-based) bitta xil validatsiya mantig'ini
takrorlagan edi -- shu sabab ikkisi orasida qoidalar sekin-asta
farqlanib qolgan (masalan 90 ta savol chegarasi faqat birida bor edi).
Endi bitta YAGONA validatsiya funksiyasi bor, ikkala chaqiruvchi ham
shundan foydalanadi va faqat natijani o'z formatiga o'girib beradi.
"""
import openpyxl

REQUIRED_COLUMNS = ["T/r", "Fan", "Savol", "A", "B", "C", "D", "To'g'ri javob", "Ball"]
LETTER_TO_INDEX = {"A": 0, "B": 1, "C": 2, "D": 3}
MAX_QUESTIONS = 90
MAX_PER_SUBJECT_BLOCK = 30


class ExcelValidationError(Exception):
    """Shablonda odam tuzatishi kerak bo'lgan xato."""


def read_and_validate_rows(file_obj_or_path, sheet_name="Savollar"):
    """Excel'ni o'qib, har bir qator uchun ODDIY (hali qaysi maqsadga
    moslanmagan) validatsiya qilingan lug'at ro'yxatini qaytaradi:

        {"row_num", "fan", "ball", "savol", "variantlar" (4 ta),
         "togri_letter"}

    Chaqiruvchi (excel_loader.py / excel_import.py) shu ro'yxatni
    o'z kerakli formatiga (togri_index yoki togri_javob, va h.k.)
    o'girib oladi.
    """
    wb = openpyxl.load_workbook(file_obj_or_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    header = [c.value for c in ws[1]]
    col_idx = {}
    for name in REQUIRED_COLUMNS:
        if name not in header:
            raise ExcelValidationError(
                f"Ustun topilmadi: '{name}'. Shablon sarlavha qatorini (1-qator) o'zgartirmang."
            )
        col_idx[name] = header.index(name)

    rows = []
    warnings = []
    fan_block_count = 0
    last_fan = None

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        def cell(name):
            return row[col_idx[name]]

        savol = cell("Savol")
        fan = cell("Fan")
        togri = cell("To'g'ri javob")
        ball = cell("Ball")
        variantlar = [cell("A"), cell("B"), cell("C"), cell("D")]

        if not savol or not fan:
            warnings.append(f"{row_num}-qator: 'Fan' yoki 'Savol' bo'sh -- o'tkazib yuborildi.")
            continue
        if not togri or str(togri).strip().upper() not in LETTER_TO_INDEX:
            raise ExcelValidationError(
                f"{row_num}-qator: 'To'g'ri javob' faqat A, B, C yoki D bo'lishi kerak "
                f"(topildi: {togri!r})."
            )
        if any(v is None or str(v).strip() == "" for v in variantlar):
            raise ExcelValidationError(f"{row_num}-qator: A/B/C/D variantlaridan biri bo'sh.")
        try:
            ball_val = float(ball)
        except (TypeError, ValueError):
            raise ExcelValidationError(f"{row_num}-qator: 'Ball' raqam bo'lishi kerak (topildi: {ball!r}).")

        fan = str(fan).strip()
        if fan != last_fan:
            fan_block_count = 0
            last_fan = fan
        fan_block_count += 1
        if fan_block_count > MAX_PER_SUBJECT_BLOCK:
            warnings.append(
                f"{row_num}-qator: '{fan}' fani {MAX_PER_SUBJECT_BLOCK} tadan ortiq savolga ega -- "
                f"joriy javoblar varag'i shablonida bitta ustunga {MAX_PER_SUBJECT_BLOCK} tagacha savol sig'adi."
            )

        rows.append({
            "row_num": row_num,
            "fan": fan,
            "ball": ball_val,
            "savol": str(savol).strip(),
            "variantlar": [str(v).strip() for v in variantlar],
            "togri_letter": str(togri).strip().upper(),
        })

    if not rows:
        raise ExcelValidationError("Faylda birorta ham to'liq savol topilmadi.")
    if len(rows) > MAX_QUESTIONS:
        raise ExcelValidationError(
            f"Jami {len(rows)} ta savol -- joriy shablon {MAX_QUESTIONS} tagacha savolni qo'llab-quvvatlaydi."
        )

    return rows, warnings