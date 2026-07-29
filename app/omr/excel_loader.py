# -*- coding: utf-8 -*-
"""
savollar_shabloni.xlsx faylini o'qib, generate_question_booklet.py kutayotgan
formatga o'tkazadi. Bu qism -- o'qituvchi yuklagan Excel bilan booklet
generatori orasidagi "ko'prik".
"""

import openpyxl

REQUIRED_COLUMNS = ["T/r", "Fan", "Savol", "A", "B", "C", "D", "To'g'ri javob", "Ball"]
LETTER_TO_INDEX = {"A": 0, "B": 1, "C": 2, "D": 3}


class TemplateError(Exception):
    """Shablonda odam tuzatishi kerak bo'lgan xato bo'lsa shu ko'tariladi."""


def load_questions_from_excel(path, sheet_name="Savollar"):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    header = [c.value for c in ws[1]]
    col_idx = {}
    for name in REQUIRED_COLUMNS:
        if name not in header:
            raise TemplateError(
                f"Ustun topilmadi: '{name}'. Shablon sarlavha qatorini (1-qator) o'zgartirmang."
            )
        col_idx[name] = header.index(name)

    questions = []
    warnings = []
    position = 0
    last_fan = None
    fan_block_count = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue  # bo'sh qator -- o'tkazib yuboriladi

        def cell(name):
            return row[col_idx[name]]

        savol = cell("Savol")
        fan = cell("Fan")
        togri = cell("To'g'ri javob")
        ball = cell("Ball")
        variantlar = [cell("A"), cell("B"), cell("C"), cell("D")]

        if not savol or not fan:
            warnings.append(f"{row_num}-qator: 'Fan' yoki 'Savol' bo'sh -- o'tkazib yuborildi.")
            continue
        if not togri or str(togri).strip().upper() not in LETTER_TO_INDEX:
            raise TemplateError(
                f"{row_num}-qator: 'To'g'ri javob' faqat A, B, C yoki D bo'lishi kerak "
                f"(topildi: {togri!r})."
            )
        if any(v is None or str(v).strip() == "" for v in variantlar):
            raise TemplateError(f"{row_num}-qator: A/B/C/D variantlaridan biri bo'sh.")
        try:
            ball_val = float(ball)
        except (TypeError, ValueError):
            raise TemplateError(f"{row_num}-qator: 'Ball' raqam bo'lishi kerak (topildi: {ball!r}).")

        position += 1
        fan = str(fan).strip()
        if fan != last_fan:
            fan_block_count = 0
            last_fan = fan
        fan_block_count += 1
        if fan_block_count > 30:
            warnings.append(
                f"{row_num}-qator: '{fan}' fani 30 tadan ortiq savolga ega -- joriy javoblar "
                f"varag'i shablonida bitta ustunga 30 tagacha savol sig'adi."
            )

        questions.append({
            "id": position,
            "fan": fan,
            "ball": f"{ball_val:g}",
            "savol": str(savol).strip(),
            "variantlar": [str(v).strip() for v in variantlar],
            "togri_index": LETTER_TO_INDEX[str(togri).strip().upper()],
        })

    if not questions:
        raise TemplateError("Faylda birorta ham to'liq savol topilmadi.")
    if len(questions) > 90:
        raise TemplateError(
            f"Jami {len(questions)} ta savol -- joriy shablon 90 tagacha savolni qo'llab-quvvatlaydi."
        )

    return questions, warnings
